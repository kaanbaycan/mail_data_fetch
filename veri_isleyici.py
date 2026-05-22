import openpyxl
import os
import glob

SOURCE_DIR = "indirilen_ekler"
OUTPUT_FILE = "Guncel_Master_Veri.xlsx"

def get_column_data(ws, col_index, start_row=6):
    """Belirli bir sütundaki verileri bir liste olarak döner."""
    data = {}
    for row in range(start_row, ws.max_row + 1):
        data[row] = ws.cell(row=row, column=col_index).value
    return data

def process_and_merge_files():
    files = glob.glob(os.path.join(SOURCE_DIR, "*.xlsx"))
    if len(files) < 2:
        print("HATA: İşlem için 2 adet dosya bulunamadı. Lütfen maillerin indiğinden emin olun.")
        return

    print(f"Dosyalar analiz ediliyor: {[os.path.basename(f) for f in files]}")

    # Dosyaları yükle
    wb1 = openpyxl.load_workbook(files[0])
    ws1 = wb1.active
    wb2 = openpyxl.load_workbook(files[1])
    ws2 = wb2.active

    # 5. satırdaki başlıkları bul
    headers1 = [cell.value for cell in ws1[5]]
    headers2 = [cell.value for cell in ws2[5]]

    try:
        col_bcsl = headers1.index("BCSL0018") + 1
        col_aazbn = headers1.index("AAZBN00") + 1
    except ValueError:
        print("HATA: Sütun isimleri (BCSL0018 veya AAZBN00) 5. satırda bulunamadı!")
        return

    # MANTIK: Hangi dosyada hangi sütun daha "güncel/dolu"?
    # X'in dosyasında BCSL0018 dolu, Y'nin dosyasında AAZBN00 dolu.
    
    def count_filled(ws, col):
        return sum(1 for row in range(6, 20) if ws.cell(row=row, column=col).value is not None)

    # Dosya 1'deki doluluk oranları
    f1_bcsl_score = count_filled(ws1, col_bcsl)
    f1_aazbn_score = count_filled(ws1, col_aazbn)
    
    # Dosya 2'deki doluluk oranları
    f2_bcsl_score = count_filled(ws2, col_bcsl)
    f2_aazbn_score = count_filled(ws2, col_aazbn)

    # Karar ver: BCSL hangi dosyadan alınacak, AAZBN hangi dosyadan?
    source_ws_bcsl = ws1 if f1_bcsl_score >= f2_bcsl_score else ws2
    source_ws_aazbn = ws1 if f1_aazbn_score > f2_aazbn_score else ws2

    print(f"BCSL0018 verisi '{source_ws_bcsl.parent.filename}' dosyasından alınıyor.")
    print(f"AAZBN00 verisi '{source_ws_aazbn.parent.filename}' dosyasından alınıyor.")

    # Yeni bir Master dosya oluştur (ws1'i şablon olarak kullan)
    # ws1 üzerinde her iki sütunu da ilgili kaynaklardan gelen verilerle güncelle
    for row in range(6, ws1.max_row + 1):
        # BCSL sütununu ilgili kaynaktan yaz
        ws1.cell(row=row, column=col_bcsl).value = source_ws_bcsl.cell(row=row, column=col_bcsl).value
        # AAZBN sütununu ilgili kaynaktan yaz
        ws1.cell(row=row, column=col_aazbn).value = source_ws_aazbn.cell(row=row, column=col_aazbn).value

    # Sonucu kaydet
    wb1.save(OUTPUT_FILE)
    print(f"\nİŞLEM BAŞARILI!")
    print(f"Kayıt yeri: {os.path.abspath(OUTPUT_FILE)}")

if __name__ == "__main__":
    process_and_merge_files()
